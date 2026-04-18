import tkinter as tk
import tkinter.filedialog as fd
from tkinter import messagebox
import xlsxwriter as xl
import openpyxl
import os
import subprocess
import sys
import datetime
import time
import threading


class App:
    def __init__(self, master):
        self.master = master
        self.master.title("Conversor SPED-Excel - por Renato - 2026")
        self.master.geometry("400x300")
        self.master.resizable(False, False)

        # Definir data limite para uso do programa
        self.end_date = datetime.date(2026, 10, 31)

        # Frame principal com padding
        main_frame = tk.Frame(self.master, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Título
        title_label = tk.Label(main_frame, text="Conversor SPED ↔ Excel", 
                              font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 20))

        # Botão para selecionar arquivo TXT
        self.select_button = tk.Button(
            main_frame, 
            text="📄 Converter SPED → Excel", 
            command=self.select_file_txt,
            width=30,
            height=2
        )
        self.select_button.pack(pady=10)

        # Botão para selecionar arquivo Excel
        self.select_excel_button = tk.Button(
            main_frame, 
            text="📊 Converter Excel → SPED", 
            command=self.select_file,
            width=30,
            height=2
        )
        self.select_excel_button.pack(pady=10)

        # Frame para checkboxes
        checkbox_frame = tk.Frame(main_frame)
        checkbox_frame.pack(pady=20)

        # Checkbutton para abrir automaticamente
        self.open_file_var = tk.BooleanVar(value=True)
        self.open_file_checkbutton = tk.Checkbutton(
            checkbox_frame, 
            text="✓ Abrir arquivo ao salvar",
            variable=self.open_file_var
        )
        self.open_file_checkbutton.pack(anchor=tk.W)

        # Checkbutton para incluir fórmula
        self.include_formula_var = tk.BooleanVar()
        self.include_formula_checkbutton = tk.Checkbutton(
            checkbox_frame, 
            text="✓ Incluir fórmula de pai",
            variable=self.include_formula_var
        )
        self.include_formula_checkbutton.pack(anchor=tk.W)

        # Rótulo de status
        self.conversion_label = tk.Label(
            main_frame, 
            text="Aguardando ação...", 
            font=("Arial", 10),
            fg="blue"
        )
        self.conversion_label.pack(pady=20)

        self.master.update()

    def select_file_txt(self):
        """Abre diálogo para selecionar arquivo TXT (SPED)"""
        # Verificar se a data limite já passou
        if datetime.date.today() > self.end_date:
            messagebox.showerror("Erro", "Data limite do programa expirou.")
            return

        filetypes = (("Arquivos TXT", "*.txt"), ("Todos os arquivos", "*.*"))
        filename = fd.askopenfilename(
            title="Selecione arquivo SPED (TXT)",
            filetypes=filetypes
        )

        if filename:
            # Executar conversão em thread separada
            threading.Thread(
                target=self.convert_txt_to_excel_threaded, 
                args=(filename,),
                daemon=True
            ).start()

    def convert_txt_to_excel_threaded(self, filename):
        """Executa conversão em thread separada"""
        try:
            self.convert_to_excel(filename)
        except Exception as e:
            self.conversion_label.config(text=f"Erro: {str(e)}", fg="red")
            messagebox.showerror("Erro", f"Erro na conversão: {str(e)}")

    def select_file(self):
        """Abre diálogo para selecionar arquivo Excel"""
        # Verificar se a data limite já passou
        if datetime.date.today() > self.end_date:
            messagebox.showerror("Erro", "Data limite do programa expirou.")
            return

        filetypes = (("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
        filename = fd.askopenfilename(
            title="Selecione arquivo Excel",
            filetypes=filetypes
        )

        if filename:
            threading.Thread(
                target=self.convert_excel_to_txt_threaded,
                args=(filename,),
                daemon=True
            ).start()

    def convert_excel_to_txt_threaded(self, filename):
        """Executa conversão Excel->TXT em thread separada"""
        try:
            self.convert_to_txt(filename)
        except Exception as e:
            self.conversion_label.config(text=f"Erro: {str(e)}", fg="red")
            messagebox.showerror("Erro", f"Erro na conversão: {str(e)}")

    def convert_to_excel(self, filename):
        """Converte arquivo TXT (SPED) para Excel"""
        start_time = time.time()
        self.conversion_label.config(text="Processamento: Aguarde...", fg="blue")
        self.master.update()

        # Validação do arquivo
        if not os.path.exists(filename):
            self.conversion_label.config(
                text="Erro: Arquivo não encontrado", 
                fg="red"
            )
            messagebox.showerror("Erro", "Arquivo não encontrado.")
            return

        # Leitura do arquivo com tratamento de erros
        try:
            with open(filename, 'r', encoding='ISO-8859-1', errors='ignore') as file:
                lines = [line.rstrip('\n') for line in file]
        except Exception as error:
            self.conversion_label.config(
                text="Erro ao abrir arquivo", 
                fg="red"
            )
            messagebox.showerror("Erro", f"Falha ao abrir o arquivo: {error}")
            return

        if not lines:
            self.conversion_label.config(
                text="Erro: Arquivo vazio", 
                fg="red"
            )
            messagebox.showwarning("Aviso", "O arquivo TXT está vazio.")
            return

        # Preparar nome do arquivo Excel
        base, _ = os.path.splitext(filename)
        excel_filename = f"{base}_v1.xlsx"

        # Verificar se arquivo já existe
        if os.path.exists(excel_filename):
            answer = messagebox.askquestion(
                "Arquivo existe",
                f"O arquivo {os.path.basename(excel_filename)} já existe.\n\n"
                "Deseja sobrescrevê-lo?"
            )
            if answer == "no":
                i = 1
                while os.path.exists(f"{base}_v{i}.xlsx"):
                    i += 1
                excel_filename = f"{base}_v{i}.xlsx"

        try:
            # Criar workbook
            workbook = xl.Workbook(excel_filename)
            bold_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D3D3D3',
                'border': 1
            })
            
            # Formato para dados
            data_format = workbook.add_format({'border': 1})
            
            worksheets = {}
            row_counters = {}
            line_numbers = {}

            # Processar linhas do arquivo
            for line_num, line in enumerate(lines, start=1):
                # Ignorar linhas que não começam com |
                if not line.startswith("|"):
                    continue

                # Dividir linha por |
                row = line.strip().split("|")
                
                # Remover último elemento vazio se existir
                if row and row[-1] == "":
                    row = row[:-1]

                # Validar linha
                if len(row) < 3:  # Mínimo: |, REG, dados
                    continue

                # Extrair nome da planilha (segundo elemento)
                sheet_name = row[1] if row[1] else "Sheet1"
                safe_sheet_name = sheet_name[:31]  # Excel tem limite de 31 caracteres

                # Criar planilha se não existir
                if safe_sheet_name not in worksheets:
                    worksheet = workbook.add_worksheet(safe_sheet_name)
                    worksheets[safe_sheet_name] = worksheet
                    row_counters[safe_sheet_name] = 1
                    line_numbers[safe_sheet_name] = []

                    # Criar cabeçalho
                    num_cols = len(row) - 1
                    header = ["ID", "REG"]
                    header.extend([f"COL_{i:02d}" for i in range(1, num_cols - 1)])
                    
                    if self.include_formula_var.get():
                        header.append("REGPAI")

                    # Escrever cabeçalho
                    for col_num, header_text in enumerate(header):
                        worksheet.write(0, col_num, header_text, bold_format)

                    # Ajustar largura das colunas
                    worksheet.set_column(0, len(header) - 1, 20)

                # Obter referência da planilha
                worksheet = worksheets[safe_sheet_name]
                current_row = row_counters[safe_sheet_name]

                # Preparar dados
                data = [line_num, row[1]]  # ID (line_num) e REG
                data.extend(row[2:-1])  # Dados intermediários

                if self.include_formula_var.get():
                    data.append("")  # Coluna para fórmula REGPAI

                # Escrever linha
                for col_num, cell_data in enumerate(data):
                    worksheet.write(current_row, col_num, cell_data, data_format)

                row_counters[safe_sheet_name] += 1
                line_numbers[safe_sheet_name].append(line_num)

            workbook.close()

            # Calcular tempo decorrido
            end_time = time.time()
            elapsed = end_time - start_time
            
            self.conversion_label.config(
                text=f"✓ Sucesso! Tempo: {elapsed:.2f}s", 
                fg="green"
            )

            # Abrir arquivo se selecionado
            if self.open_file_var.get():
                self.open_file(excel_filename)

            messagebox.showinfo(
                "Sucesso",
                f"Arquivo convertido com sucesso!\n\n"
                f"Salvo em: {os.path.basename(excel_filename)}\n"
                f"Tempo: {elapsed:.2f}s"
            )

        except Exception as e:
            self.conversion_label.config(
                text=f"Erro na conversão: {str(e)}", 
                fg="red"
            )
            messagebox.showerror("Erro", f"Erro ao criar arquivo Excel: {str(e)}")

    def convert_to_txt(self, filename):
        """Converte arquivo Excel para TXT (SPED)"""
        start_time = time.time()
        self.conversion_label.config(text="Processamento: Aguarde...", fg="blue")
        self.master.update()

        # Validação
        if not os.path.exists(filename):
            self.conversion_label.config(
                text="Erro: Arquivo não encontrado", 
                fg="red"
            )
            messagebox.showerror("Erro", "Arquivo Excel não encontrado.")
            return

        try:
            # Carregar workbook
            wb = openpyxl.load_workbook(filename, read_only=True)
            
            if not wb.sheetnames:
                self.conversion_label.config(
                    text="Erro: Nenhuma planilha encontrada", 
                    fg="red"
                )
                messagebox.showerror("Erro", "O arquivo Excel não contém planilhas.")
                return

            # Preparar nome do arquivo TXT
            base, _ = os.path.splitext(filename)
            txt_filename = f"{base}.txt"

            # Coletar linhas de todas as planilhas
            all_rows = []

            for sheet_name in wb.sheetnames:
                worksheet = wb[sheet_name]
                
                # Começar a partir da linha 2 (ignorar cabeçalho)
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
                    if not row or all(cell is None for cell in row):
                        continue

                    # Converter None em string vazia
                    row = ['' if cell is None else str(cell).strip() for cell in row]

                    all_rows.append({
                        'id': row[0] if len(row) > 0 else '',
                        'data': row,
                        'sheet': sheet_name
                    })

            if not all_rows:
                self.conversion_label.config(
                    text="Erro: Nenhum dado encontrado", 
                    fg="red"
                )
                messagebox.showwarning("Aviso", "Nenhum dado foi encontrado no Excel.")
                return

            # Ordenar por ID (primeiro elemento)
            try:
                all_rows.sort(key=lambda x: int(x['id']) if x['id'].isdigit() else float('inf'))
            except (ValueError, IndexError):
                # Se não conseguir ordenar por ID, manter ordem original
                pass

            # Escrever arquivo TXT
            with open(txt_filename, 'w', encoding='ISO-8859-1') as txt_file:
                for row_data in all_rows:
                    # Construir linha com formato SPED
                    # Formato: |REG|COL1|COL2|...|
                    row = row_data['data']
                    
                    # Ignorar primeira coluna (ID) e última coluna (se vazia/fórmula)
                    line_content = []
                    
                    if len(row) > 1:
                        # Adicionar REG (segunda coluna)
                        line_content.append(str(row[1]))
                        
                        # Adicionar dados intermediários
                        for cell in row[2:-1]:
                            line_content.append(str(cell))
                    
                    # Montar linha com pipes
                    line = "|" + "|".join(line_content) + "|\n"
                    txt_file.write(line)

            wb.close()

            # Calcular tempo decorrido
            end_time = time.time()
            elapsed = end_time - start_time

            self.conversion_label.config(
                text=f"✓ Sucesso! Tempo: {elapsed:.2f}s", 
                fg="green"
            )

            # Abrir arquivo se selecionado
            if self.open_file_var.get():
                self.open_file(txt_filename)

            messagebox.showinfo(
                "Sucesso",
                f"Arquivo convertido com sucesso!\n\n"
                f"Salvo em: {os.path.basename(txt_filename)}\n"
                f"Tempo: {elapsed:.2f}s"
            )

        except Exception as e:
            self.conversion_label.config(
                text=f"Erro na conversão: {str(e)}", 
                fg="red"
            )
            messagebox.showerror("Erro", f"Erro ao converter Excel: {str(e)}")

    @staticmethod
    def open_file(filepath):
        """Abre arquivo com aplicação padrão do sistema"""
        try:
            if os.name == "nt":  # Windows
                os.startfile(filepath)
            elif sys.platform == "darwin":  # macOS
                subprocess.run(["open", filepath], check=False)
            else:  # Linux
                subprocess.run(["xdg-open", filepath], check=False)
        except Exception as error:
            print(f"Não foi possível abrir o arquivo: {error}")


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()